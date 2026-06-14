"""Generate a Monthly Expense Tracker .xlsx from an OCBC FRANK CSV export.

Reads TransactionHistory_*.csv, parses the OCBC export, auto-categorizes each
spend row using merchant-text rules, and writes MonthlyExpenseTracker.xlsx with
three tabs (Dashboard, Transactions, Setup) per instructions.md.

The Dashboard has a Month selector so all figures can be viewed per-month or for
"All" months. Spend rows that don't match a known merchant rule default to
Variable Wants -> Shopping (flagged in Notes so they're easy to retag).
"""

import glob
import re
import sys
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import PatternFillProperties  # noqa: F401
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.workbook.defined_name import DefinedName

# --- Spending taxonomy ---------------------------------------------------------
# Money is split by transaction TYPE. "Fixed Needs"/"Variable Wants" are spending;
# "Transfer" is money moved (savings/investment + person-to-person) and is EXCLUDED
# from spending. Savings is an outcome (Income - Spending), evaluated as a share of
# income against the 50/30/20 rule.
CATEGORIES = {
    "Fixed Needs": [
        "Accommodation/Rent", "Transport", "Insurance",
        "Basic Groceries", "Utilities",
    ],
    "Variable Wants": [
        "Dining Out/Cafes", "Entertainment/Hobbies", "Subscriptions",
        "Shopping", "Travel",
    ],
    "Transfer": [
        "Savings / Investment", "Personal Transfer", "Reimbursement",
        "Other Transfer",
    ],
}
SPENDING_PILLARS = ["Fixed Needs", "Variable Wants"]
TRANSFER_PILLAR = "Transfer"

# Budget buckets for the 50/30/20 view (share of income).
BUDGET_TARGETS = {"Needs": 0.50, "Wants": 0.30, "Savings": 0.20}

DEFAULT_CATEGORY = ("Variable Wants", "Shopping")  # fallback for unknown merchants
TRANSFER_DEFAULT = ("Transfer", "Personal Transfer")
SAVINGS_DEFAULT = ("Transfer", "Savings / Investment")

# --- Merchant -> (Pillar, Sub-Category) rules ---------------------------------
# Order matters: first matching keyword wins. Matched case-insensitively against
# the whitespace-normalized description text.
RULES = [
    (r"BUS/MRT|CAUSEWAYLINK|GOPAY|GOJEK|GRAB|COMFORTDELGRO|SMRT|TRANSITLINK",
     ("Fixed Needs", "Transport")),
    (r"SHENG SIONG|NTUC|FAIRPRICE|NTUC FP|GIANT|COLD STORAGE|GOURMET PARADISE|"
     r"7-ELEVEN|7 ELEVEN|KK MART|KK SUPER|99 SPEEDMART|CHEERS|PRIME SUPER",
     ("Fixed Needs", "Basic Groceries")),
    (r"SP SERVICES|SP SVCS|SINGTEL|STARHUB|M1 |MYREPUBLIC|CIRCLES\.LIFE|"
     r"SINGTEL PREPAID|PUB ",
     ("Fixed Needs", "Utilities")),
    (r"INSURANCE|AIA|PRUDENTIAL|GREAT EASTERN|NTUC INCOME|MANULIFE",
     ("Fixed Needs", "Insurance")),
    (r"RENT|RENTAL|ACCOMMODATION|HOSTEL|LANDLORD",
     ("Fixed Needs", "Accommodation/Rent")),
    (r"STRIPE|NETFLIX|SPOTIFY|YOUTUBE|ICLOUD|APPLE\.COM|GOOGLE \*|"
     r"OPENAI|CHATGPT|ADOBE|MICROSOFT|AMAZON PRIME|DISNEY",
     ("Variable Wants", "Subscriptions")),
    (r"BURGER KING|MCDONALD|KFC|STARBUCKS|COFFEE|CAFE|KOPITIAM|FOODPANDA|"
     r"DELIVEROO|ASIAN ROTISSERIE|IRON CHEF|A KITCHEN|CHOCOLICIOUS|RESTAURANT|"
     r"EATERY|F&B|BAKERY|TOAST|BUBBLE|DESSERT|HAWKER|FOOD|DELIGHT|TECHNO EDGE|"
     r"KITCHEN|HWANG",
     ("Variable Wants", "Dining Out/Cafes")),
    (r"SHOPEE|LAZADA|UNIQLO|H&M|MALL|WATSON|GUARDIAN|DAISO|MUJI|DECATHLON|"
     r"POPULAR|CHALLENGER|COURTS",
     ("Variable Wants", "Shopping")),
    (r"CATHAY|GOLDEN VILLAGE|GV |SHAW|CINEMA|KARAOKE|KTV|ARCADE|STEAM ",
     ("Variable Wants", "Entertainment/Hobbies")),
    (r"AIRLINE|AIRASIA|SCOOT|SINGAPORE AIRLINES|HOTEL|AGODA|BOOKING\.COM|"
     r"EXPEDIA|KLOOK|AIRBNB",
     ("Variable Wants", "Travel")),
]

INCOME_KEYWORDS = re.compile(r"GIRO - SALARY|\bSALARY\b|INFI\s*NEON|TECHNOLOG\s*SALA", re.I)

# Investment / savings platforms — money here is a transfer, not spending.
INVESTMENT_RE = re.compile(
    r"SYFE|ENDOWUS|STASHAWAY|FSMONE|FUNDSUPERMART|MOOMOO|FUTU|TIGER BROKERS|"
    r"INTERACTIVE BROKERS|\bIBKR\b|WEBULL|SAXO|POEMS|DBS INVEST|REGULAR SAVINGS|"
    r"FIXED DEPOSIT|\bSSB\b|SINGAPORE SAVINGS BOND|CPF|SRS|GIGANTIQ|SINGLIFE",
    re.I,
)


def detect_transfer(desc: str):
    """Return (pillar, sub) if the row is a transfer (money moved, not spent)."""
    if INVESTMENT_RE.search(desc):
        return SAVINGS_DEFAULT
    d = desc.upper()
    # Person-to-person: mobile PayNow / "Transfer - Mobile" / fund transfer to a
    # named individual. UEN payments are businesses -> treated as spending.
    looks_p2p = (
        "PAYNOW-MOBILE" in d
        or re.search(r"TRANSFER\s*-\s*MOBILE", d)
        or (
            re.search(r"FAST PAYMENT|FUND TRANSFER|PAYMENT/TRANSFER", d)
            and re.search(r"\bTO\s+[A-Z]", d)
            and "PAYNOW-UEN" not in d
        )
    )
    if looks_p2p:
        return TRANSFER_DEFAULT
    return None


def find_csv() -> str:
    files = sorted(glob.glob("TransactionHistory_*.csv"))
    if not files:
        sys.exit("No TransactionHistory_*.csv found in current directory.")
    return files[-1]


def normalize(text: str) -> str:
    return re.sub(r"\s+", " ", str(text)).strip()


def categorize(desc: str):
    """Return (pillar, sub, kind) where kind is 'rule', 'transfer', or 'default'."""
    for pattern, (pillar, sub) in RULES:
        if re.search(pattern, desc, re.I):
            return pillar, sub, "rule"
    transfer = detect_transfer(desc)
    if transfer:
        return transfer[0], transfer[1], "transfer"
    return DEFAULT_CATEGORY[0], DEFAULT_CATEGORY[1], "default"


def parse_csv(path: str) -> pd.DataFrame:
    header_idx = None
    with open(path, "r", encoding="utf-8-sig") as fh:
        for i, line in enumerate(fh):
            if line.strip().lower().startswith("transaction date"):
                header_idx = i
                break
    if header_idx is None:
        sys.exit("Could not locate the 'Transaction date' header row.")
    # skip_blank_lines=False keeps pandas' row indexing aligned with the physical
    # line numbers used to find header_idx; otherwise the blank preamble line is
    # skipped and the first data row is mistakenly consumed as the header.
    df = pd.read_csv(
        path,
        header=header_idx,
        dtype=str,
        keep_default_na=False,
        skip_blank_lines=False,
    )
    df.columns = ["Date", "ValueDate", "Description", "Withdrawals", "Deposits"][: len(df.columns)]
    df = df[df["Date"].str.strip() != ""]
    return df


def to_amount(val: str) -> float:
    val = str(val).replace(",", "").strip()
    if val == "":
        return 0.0
    try:
        return float(val)
    except ValueError:
        return 0.0


def to_date(val: str):
    try:
        return datetime.strptime(val.strip(), "%d/%m/%Y")
    except ValueError:
        return None


def main():
    csv_path = find_csv()
    df = parse_csv(csv_path)

    rows = []
    income_by_month = {}
    for _, r in df.iterrows():
        desc = normalize(r["Description"])
        dt = to_date(r["Date"])
        withdrawal = to_amount(r["Withdrawals"])
        deposit = to_amount(r["Deposits"])
        month = dt.strftime("%Y-%m") if dt else ""
        if deposit > 0 and withdrawal == 0:
            if INCOME_KEYWORDS.search(desc) and month:
                income_by_month[month] = income_by_month.get(month, 0.0) + deposit
            continue
        if withdrawal == 0:
            continue
        pillar, sub, kind = categorize(desc)
        if kind == "transfer":
            note = "transfer - excluded from spending"
        elif kind == "default":
            note = "auto-default (Shopping) - retag if needed"
        else:
            note = ""
        rows.append({
            "Date": dt,
            "Description": desc,
            "Amount": round(withdrawal, 2),
            "Main Pillar": pillar,
            "Sub-Category": sub,
            "Notes": note,
            "_month": month,
        })

    tx = pd.DataFrame(rows)
    build_workbook(tx, income_by_month, csv_path)
    n_default = (tx["Main Pillar"].isin(SPENDING_PILLARS) & (tx["Notes"].str.startswith("auto-default"))).sum()
    n_transfer = (tx["Main Pillar"] == TRANSFER_PILLAR).sum()
    n_spend = tx["Main Pillar"].isin(SPENDING_PILLARS).sum()
    print(f"Parsed {len(tx)} outflow rows from {csv_path}")
    print(f"Spending rows: {n_spend} (defaulted to Shopping: {n_default}); transfers: {n_transfer}")
    print(f"Months present: {sorted(m for m in tx['_month'].unique() if m)}")
    print(f"Detected salary income by month: {income_by_month}")
    print("Wrote MonthlyExpenseTracker.xlsx")


# ------------------------------------------------------------------ styling ---
# Apple-inspired palette: ink, Action Blue, iOS orange/green.
CHARCOAL = "1D1D1F"   # ink (headers/titles)
GREEN = "34C759"      # savings / positive
SAFFRON = "FF9500"    # wants / warning
ORANGE = "FF9500"
CORAL = "FF3B30"      # spent / negative
BLUE = "0066CC"       # needs / accent
PILLAR_COLORS = [BLUE, SAFFRON]          # spending pie: Fixed Needs / Variable Wants
BUDGET_COLORS = [BLUE, SAFFRON, GREEN]   # budget bars: Needs / Wants / Savings

THIN = Side(style="thin", color="DCE3EA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor=CHARCOAL)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=18, color=CHARCOAL)
SUBTITLE_FONT = Font(bold=True, size=12, color=GREEN)
LABEL_FONT = Font(bold=True, color="3D4852")
BAND_FILL = PatternFill("solid", fgColor="F1F5F7")
CARD_FILL = PatternFill("solid", fgColor="FFFFFF")
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
RED_FILL = PatternFill("solid", fgColor="FAD9CF")
GREEN_FILL = PatternFill("solid", fgColor="D5EEE9")


def style_header_row(ws, row, ncols, start=1):
    for c in range(start, start + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = CENTER


def hide_gridlines(ws):
    ws.sheet_view.showGridLines = False


def build_workbook(tx: pd.DataFrame, income_by_month: dict, csv_path: str):
    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_tx = wb.create_sheet("Transactions")
    ws_set = wb.create_sheet("Setup")

    months = sorted(m for m in tx["_month"].unique() if m)
    pillars = list(CATEGORIES.keys())

    for ws in (ws_dash, ws_tx, ws_set):
        hide_gridlines(ws)
    ws_dash.sheet_properties.tabColor = CHARCOAL
    ws_tx.sheet_properties.tabColor = GREEN
    ws_set.sheet_properties.tabColor = SAFFRON

    _build_setup(wb, ws_set, months, income_by_month, pillars)
    last_tx = _build_transactions(ws_tx, tx, pillars)
    _build_dashboard(ws_dash, pillars, months, last_tx)

    wb.save("MonthlyExpenseTracker.xlsx")


def _build_setup(wb, ws, months, income_by_month, pillars):
    ws["A1"] = "Setup - Categories, Targets & Income"
    ws["A1"].font = TITLE_FONT

    ws["A3"] = "Pillars"
    ws["A3"].font, ws["A3"].fill = HEADER_FONT, HEADER_FILL
    for i, p in enumerate(pillars):
        ws.cell(row=4 + i, column=1, value=p)

    start_col = 3
    for c, p in enumerate(pillars):
        col = start_col + c
        cl = get_column_letter(col)
        hc = ws.cell(row=3, column=col, value=p)
        hc.font, hc.fill = HEADER_FONT, HEADER_FILL
        subs = CATEGORIES[p]
        for i, s in enumerate(subs):
            ws.cell(row=4 + i, column=col, value=s)
        rng = f"Setup!${cl}$4:${cl}${3 + len(subs)}"
        wb.defined_names.add(DefinedName(p.replace(" ", "_"), attr_text=rng))
    wb.defined_names.add(
        DefinedName("Pillars", attr_text=f"Setup!$A$4:$A${3 + len(pillars)}"))

    ws["A11"] = "Baseline targets (% of income)"
    ws["A11"].font = LABEL_FONT
    for i, (bucket, val) in enumerate(BUDGET_TARGETS.items()):
        ws.cell(row=12 + i, column=1, value=bucket)
        tc = ws.cell(row=12 + i, column=2, value=val)
        tc.number_format = "0%"

    # Month options for the dropdown: "All" + each month present.
    ws["E11"] = "Month options"
    ws["E11"].font = LABEL_FONT
    opts = ["All"] + months
    for i, m in enumerate(opts):
        ws.cell(row=12 + i, column=5, value=m)
    wb.defined_names.add(
        DefinedName("MonthOptions",
                    attr_text=f"Setup!$E$12:$E${11 + len(opts)}"))

    # Income-by-month table (from detected GIRO-SALARY deposits).
    ws["G11"] = "Month"
    ws["H11"] = "Salary Income"
    ws["G11"].font = ws["H11"].font = LABEL_FONT
    for i, m in enumerate(months):
        ws.cell(row=12 + i, column=7, value=m)
        ic = ws.cell(row=12 + i, column=8, value=round(income_by_month.get(m, 0.0), 2))
        ic.number_format = "#,##0.00"
    last_inc = 11 + len(months)
    wb.defined_names.add(
        DefinedName("IncomeMonths", attr_text=f"Setup!$G$12:$G${last_inc}"))
    wb.defined_names.add(
        DefinedName("IncomeAmounts", attr_text=f"Setup!$H$12:$H${last_inc}"))

    # Selected-month date window, derived from the Dashboard month selector.
    ws["A20"] = "Selected window (auto)"
    ws["A20"].font = Font(italic=True, size=9, color="999999")
    ws["A21"] = "start"
    ws["B21"] = ('=IF(Dashboard!$B$3="All",DATE(1900,1,1),'
                 'DATE(VALUE(LEFT(Dashboard!$B$3,4)),VALUE(RIGHT(Dashboard!$B$3,2)),1))')
    ws["A22"] = "end"
    ws["B22"] = '=IF(Dashboard!$B$3="All",DATE(2999,1,1),EOMONTH($B$21,0))'
    for c in ("B21", "B22"):
        ws[c].number_format = "DD/MM/YYYY"
    wb.defined_names.add(DefinedName("StartDate", attr_text="Setup!$B$21"))
    wb.defined_names.add(DefinedName("EndDate", attr_text="Setup!$B$22"))

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 20


def _build_transactions(ws, tx, pillars):
    headers = ["Date", "Description", "Amount", "Main Pillar", "Sub-Category", "Notes"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    for ri, (_, r) in enumerate(tx.iterrows(), start=2):
        dc = ws.cell(row=ri, column=1, value=r["Date"])
        dc.number_format = "DD/MM/YYYY"
        ws.cell(row=ri, column=2, value=r["Description"])
        ac = ws.cell(row=ri, column=3, value=r["Amount"])
        ac.number_format = "#,##0.00"
        ws.cell(row=ri, column=4, value=r["Main Pillar"])
        ws.cell(row=ri, column=5, value=r["Sub-Category"])
        ws.cell(row=ri, column=6, value=r["Notes"])
        if ri % 2 == 0:
            for c in range(1, 7):
                ws.cell(row=ri, column=c).fill = BAND_FILL

    last = len(tx) + 1
    for c, w in enumerate((12, 54, 12, 16, 22, 34), start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"

    dv_max = max(last, 1000)
    dv_pillar = DataValidation(type="list", formula1="=Pillars", allow_blank=True)
    ws.add_data_validation(dv_pillar)
    dv_pillar.add(f"D2:D{dv_max}")
    dv_sub = DataValidation(
        type="list", formula1='=INDIRECT(SUBSTITUTE($D2," ","_"))', allow_blank=True)
    ws.add_data_validation(dv_sub)
    dv_sub.add(f"E2:E{dv_max}")
    return dv_max


def _build_dashboard(ws, pillars, months, dv_max):
    amt = f"Transactions!$C$2:$C${dv_max}"
    pil = f"Transactions!$D$2:$D${dv_max}"
    sub = f"Transactions!$E$2:$E${dv_max}"
    dat = f"Transactions!$A$2:$A${dv_max}"

    ws["A1"] = "Monthly Expense Tracker"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 30

    # --- Month selector ---
    ws["A3"] = "Month"
    ws["A3"].font = LABEL_FONT
    ws["A3"].alignment = RIGHT
    sel = ws["B3"]
    sel.value = months[-1] if months else "All"
    sel.fill = PatternFill("solid", fgColor=SAFFRON)
    sel.font = Font(bold=True, color=CHARCOAL)
    sel.alignment = CENTER
    sel.border = BORDER
    dv_month = DataValidation(type="list", formula1="=MonthOptions", allow_blank=False)
    ws.add_data_validation(dv_month)
    dv_month.add("B3")

    drange = f'{dat},">="&StartDate,{dat},"<="&EndDate'

    # --- Summary cards (spending excludes transfers; savings is an outcome) ---
    needs_f = f'SUMIFS({amt},{pil},"Fixed Needs",{drange})'
    wants_f = f'SUMIFS({amt},{pil},"Variable Wants",{drange})'
    cards = [
        ("Total Income", '=IF($B$3="All",SUM(IncomeAmounts),'
                         'SUMIFS(IncomeAmounts,IncomeMonths,$B$3))', BLUE, "money"),
        ("Total Spent", f"={needs_f}+{wants_f}", CORAL, "money"),
        ("Saved (Income - Spent)", "=B5-B6", GREEN, "money"),
        ("Savings Rate", "=IF(B5=0,0,B7/B5)", CHARCOAL, "pct"),
    ]
    for i, (label, formula, color, kind) in enumerate(cards):
        r = 5 + i
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = Font(bold=True, color="FFFFFF")
        lc.fill = PatternFill("solid", fgColor=color)
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        lc.border = BORDER
        vc = ws.cell(row=r, column=2, value=formula)
        vc.number_format = "0.0%" if kind == "pct" else "#,##0.00"
        vc.font = Font(bold=True, size=12, color=color)
        vc.alignment = RIGHT
        vc.border = BORDER
        ws.row_dimensions[r].height = 22

    # --- 50/30/20 budget table (share of income) ---
    hr = 10
    ws.cell(row=hr, column=1, value="50 / 30 / 20 - share of income").font = SUBTITLE_FONT
    hr += 1
    for c, h in enumerate(["Bucket", "Amount", "Actual %", "Target %", "Status"], 1):
        ws.cell(row=hr, column=c, value=h)
    style_header_row(ws, hr, 5)
    first = hr + 1
    budget = [
        ("Needs", f"={needs_f}", BUDGET_TARGETS["Needs"], "<="),
        ("Wants", f"={wants_f}", BUDGET_TARGETS["Wants"], "<="),
        ("Savings", "=$B$7", BUDGET_TARGETS["Savings"], ">="),
    ]
    for i, (bucket, amount_f, target, op) in enumerate(budget):
        row = first + i
        ws.cell(row=row, column=1, value=bucket)
        sp = ws.cell(row=row, column=2, value=amount_f)
        sp.number_format = "#,##0.00"
        sp.alignment = RIGHT
        pct = ws.cell(row=row, column=3, value=f"=IF($B$5=0,0,B{row}/$B$5)")
        pct.number_format = "0.0%"
        pct.alignment = CENTER
        tg = ws.cell(row=row, column=4, value=target)
        tg.number_format = "0%"
        tg.alignment = CENTER
        if op == ">=":
            st = f'=IF(C{row}>=D{row},"On track","Below target")'
        else:
            st = f'=IF(C{row}<=D{row},"On track","Over budget")'
        stc = ws.cell(row=row, column=5, value=st)
        stc.alignment = CENTER
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = BORDER
        if i % 2 == 1:
            for c in range(1, 6):
                ws.cell(row=row, column=c).fill = BAND_FILL
    last = first + len(budget) - 1

    for word in ("Over budget", "Below target"):
        ws.conditional_formatting.add(
            f"E{first}:E{last}",
            FormulaRule(formula=[f'$E{first}="{word}"'], fill=RED_FILL))
    ws.conditional_formatting.add(
        f"E{first}:E{last}",
        FormulaRule(formula=[f'$E{first}="On track"'], fill=GREEN_FILL))

    # Transfers line (excluded from spending).
    trow = last + 1
    ws.cell(row=trow, column=1, value="Transfers (excluded)").font = Font(italic=True, size=9, color="999999")
    tvc = ws.cell(row=trow, column=2, value=f'=SUMIFS({amt},{pil},"Transfer",{drange})')
    tvc.number_format = "#,##0.00"
    tvc.alignment = RIGHT
    tvc.font = Font(italic=True, size=9, color="999999")

    # --- Sub-category breakdown table ---
    sr = last + 3
    ws.cell(row=sr, column=1, value="Sub-category breakdown").font = SUBTITLE_FONT
    sr += 1
    for c, h in enumerate(["Sub-Category", "Pillar", "Spent"], 1):
        ws.cell(row=sr, column=c, value=h)
    style_header_row(ws, sr, 3)
    sub_first = sr + 1
    row = sub_first
    for p in SPENDING_PILLARS:
        for s in CATEGORIES[p]:
            ws.cell(row=row, column=1, value=s)
            ws.cell(row=row, column=2, value=p)
            sc = ws.cell(row=row, column=3, value=f'=SUMIFS({amt},{sub},"{s}",{drange})')
            sc.number_format = "#,##0.00"
            sc.alignment = RIGHT
            for c in range(1, 4):
                ws.cell(row=row, column=c).border = BORDER
            if (row - sub_first) % 2 == 1:
                for c in range(1, 4):
                    ws.cell(row=row, column=c).fill = BAND_FILL
            row += 1
    sub_last = row - 1

    # ---------------------------------------------------------------- charts ---
    # 1) Pie: spending split (Needs vs Wants), labels outside with leader lines.
    pie = PieChart()
    pie.title = "Spending: Needs vs Wants"
    pie.height, pie.width = 8.5, 14
    pie_last = first + 1  # Needs + Wants rows only (exclude Savings)
    pdata = Reference(ws, min_col=2, min_row=hr, max_row=pie_last)  # incl header for title
    pcats = Reference(ws, min_col=1, min_row=first, max_row=pie_last)
    pie.add_data(pdata, titles_from_data=True)
    pie.set_categories(pcats)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.numFmt = "0%"
    pie.dataLabels.dLblPos = "outEnd"
    pie.dataLabels.showLeaderLines = True
    for idx, color in enumerate(PILLAR_COLORS):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties = GraphicalProperties(solidFill=color)
        pie.series[0].data_points.append(pt)
    pie.legend.position = "b"
    ws.add_chart(pie, "G3")

    # 2) Column: Actual % vs Target % (share of income) for Needs/Wants/Savings.
    bar = BarChart()
    bar.type = "col"
    bar.title = "Actual vs Target (% of income)"
    bar.height, bar.width = 8.5, 14
    bdata = Reference(ws, min_col=3, max_col=4, min_row=hr, max_row=last)
    bcats = Reference(ws, min_col=1, min_row=first, max_row=last)
    bar.add_data(bdata, titles_from_data=True)
    bar.set_categories(bcats)
    bar.y_axis.numFmt = "0%"
    bar.y_axis.scaling.min = 0
    bar.y_axis.scaling.max = 1
    bar.gapWidth = 80
    bar.overlap = -20
    bar.dataLabels = DataLabelList()
    bar.dataLabels.showVal = True
    bar.dataLabels.numFmt = "0%"
    bar.dataLabels.dLblPos = "outEnd"
    bar.series[0].graphicalProperties = GraphicalProperties(solidFill=BLUE)
    bar.series[1].graphicalProperties = GraphicalProperties(solidFill="D2D2D7")
    bar.legend.position = "b"
    ws.add_chart(bar, "G21")

    # 3) Horizontal bar: spend by sub-category, WITH value labels at the bar ends.
    hbar = BarChart()
    hbar.type = "bar"
    hbar.title = "Spend by Sub-Category"
    hbar.height, hbar.width = 12, 14
    hdata = Reference(ws, min_col=3, min_row=sr, max_row=sub_last)
    hcats = Reference(ws, min_col=1, min_row=sub_first, max_row=sub_last)
    hbar.add_data(hdata, titles_from_data=True)
    hbar.set_categories(hcats)
    hbar.legend = None
    hbar.gapWidth = 40
    hbar.x_axis.delete = True  # hide the value axis; labels carry the numbers
    hbar.dataLabels = DataLabelList()
    hbar.dataLabels.showVal = True
    hbar.dataLabels.numFmt = "#,##0"
    hbar.dataLabels.dLblPos = "outEnd"
    hbar.series[0].graphicalProperties = GraphicalProperties(solidFill=BLUE)
    ws.add_chart(hbar, "G39")

    note = sub_last + 2
    ws.cell(row=note, column=1, value=(
        "Spending excludes transfers (savings/investment + person-to-person), which are "
        "money moved, not spent. Savings = Income - Spending. Income is auto-detected from "
        "monthly salary deposits (Setup tab). Rows flagged 'auto-default' were uncategorized "
        "and placed under Shopping; retag them for accurate splits."))
    ws.cell(row=note, column=1).font = Font(italic=True, size=9, color="999999")

    ws.column_dimensions["A"].width = 24
    for col in ("B", "C", "D", "E"):
        ws.column_dimensions[col].width = 14
    ws.freeze_panes = "A2"


if __name__ == "__main__":
    main()
